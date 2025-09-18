import streamlit as st
import pandas as pd
import random
from collections import defaultdict
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="审核编辑排班工具", layout="wide")
st.title("📅 审核编辑团队自助排班工具")

uploaded_file = st.file_uploader("上传编辑名单（CSV，列名：姓名）", type="csv")
previous_schedule_file = st.file_uploader("（可选）上传上月排班统计表（CSV，列名：姓名, 早班, 晚班, 常规班,周末值班）", type="csv")
past_weekend_file = st.file_uploader("（可选）上周历史周末值班表（CSV 或 Excel，列名：姓名, 周次）", type=["csv", "xlsx", "xls"])

st.sidebar.header("排班设定")
start_date = st.sidebar.date_input("起始日期", value=datetime.today())
weeks_to_schedule = st.sidebar.number_input("排班周数", min_value=1, max_value=12, value=4)
morning_count = st.sidebar.number_input("每周早班人数", min_value=1, max_value=30, value=10)
evening_count = st.sidebar.number_input("每周晚班人人", min_value=1, max_value=30, value=10)
weekend_count = st.sidebar.number_input("每周周末值班人数", min_value=1, max_value=30, value=5)

if uploaded_file:
    df = pd.read_csv(uploaded_file)
    editors = df['姓名'].tolist()
    st.sidebar.markdown("### ⛔ 排除班次设定")
    exclude_morning = st.sidebar.multiselect("不排早班的人员", editors)
    exclude_evening = st.sidebar.multiselect("不排晚班的人员", editors)
    exclude_weekend = st.sidebar.multiselect("不排周末值班的人员", editors)

    if st.button("🔄 生成排班表"):
        if morning_count + evening_count > 30:
            st.error("早班与晚班总人数不可超过 30，请重新设定。")
        elif len(editors) < 30:
            st.error("编辑人数必须为 30 人，才能确保三班不重叠。")
        else:
            past_weekend_weeks = defaultdict(set)
            if past_weekend_file:
                if past_weekend_file.name.endswith('.csv'):
                    past_df = pd.read_csv(past_weekend_file)
                elif past_weekend_file.name.endswith('.xlsx'):
                    past_df = pd.read_excel(past_weekend_file, engine="openpyxl")
                elif past_weekend_file.name.endswith('.xls'):
                    past_df = pd.read_excel(past_weekend_file, engine="xlrd")
                else:
                    st.error("请上传 .csv、.xlsx 或 .xls 格式的历史周末值班表")
                    st.stop()
                for _, row in past_df.iterrows():
                    name = row['姓名']
                    week_num = int(row['周次'])
                    past_weekend_weeks[name].add(week_num)

            shift_count = defaultdict(lambda: {"早班": 0, "晚班": 0, "常规班": 0, "周末值班": 0})
            shift_weeks = defaultdict(lambda: defaultdict(set))
            schedule = defaultdict(lambda: defaultdict(list))

            if previous_schedule_file:
                prev_df = pd.read_csv(previous_schedule_file)
                for _, row in prev_df.iterrows():
                    name = row['姓名']
                    shift_count[name]["早班"] = int(row.get("早班", 0))
                    shift_count[name]["晚班"] = int(row.get("晚班", 0))
                    shift_count[name]["常规班"] = int(row.get("常规班", 0))

            def is_eligible_for_weekend(name, current_week):
                return all(abs(current_week - w) >= 2 for w in past_weekend_weeks[name])

            shuffled_editors = editors.copy()
            random.shuffle(shuffled_editors)

            for week in range(weeks_to_schedule):
                week_start = start_date + timedelta(weeks=week)
                weekdays = [week_start + timedelta(days=i) for i in range(5)]
                weekend = [week_start + timedelta(days=i) for i in range(5, 7)]

                def recent_weeks(e, shift_type):
                    return len([w for w in shift_weeks[e][shift_type] if week - w <= 3])

                def get_sorted_candidates(shift_type, exclude_set):
                    candidates = [e for e in editors if e not in exclude_set]
                    if shift_type == "早班":
                        candidates = [e for e in candidates if e not in exclude_morning]
                    elif shift_type == "晚班":
                        candidates = [e for e in candidates if e not in exclude_evening]
                    return sorted(candidates, key=lambda x: (shift_count[x][shift_type], sum(shift_count[x].values()), recent_weeks(x, shift_type)))

                used_this_week = set()

                # 统计上周早班、晚班、周末值班名单
                if week > 0:
                    last_week_start = start_date + timedelta(weeks=week-1)
                    last_week_days = [last_week_start + timedelta(days=i) for i in range(5)]
                    last_weekend = [last_week_start + timedelta(days=i) for i in range(5, 7)]
                    last_week_morning = set()
                    last_week_evening = set()
                    last_weekend_set = set()
                    for d in last_week_days:
                        day_str = d.strftime("%Y-%m-%d")
                        if "早班" in schedule[day_str]:
                            last_week_morning.update(schedule[day_str]["早班"])
                        if "晚班" in schedule[day_str]:
                            last_week_evening.update(schedule[day_str]["晚班"])
                    for d in last_weekend:
                        day_str = d.strftime("%Y-%m-%d")
                        if "周末值班" in schedule[day_str]:
                            last_weekend_set.update(schedule[day_str]["周末值班"])
                else:
                    last_week_morning = set()
                    last_week_evening = set()
                    last_weekend_set = set()

                # 分配早班，禁止连续两周早班，且上周末值班人员最多2人
                morning_candidates = get_sorted_candidates("早班", used_this_week)
                # 先排除上周早班
                morning_candidates = [e for e in morning_candidates if e not in last_week_morning]
                # 优先不安排上周末值班人员
                morning_final = []
                morning_last_weekend_count = 0
                for e in morning_candidates:
                    if e in last_weekend_set:
                        if morning_last_weekend_count < 2:
                            morning_final.append(e)
                            morning_last_weekend_count += 1
                    else:
                        morning_final.append(e)
                    if len(morning_final) == morning_count:
                        break
                morning = morning_final
                used_this_week.update(morning)

                # 分配晚班，禁止连续两周晚班，且上周末值班人员最多2人
                evening_candidates = get_sorted_candidates("晚班", used_this_week)
                evening_candidates = [e for e in evening_candidates if e not in last_week_evening]
                evening_final = []
                evening_last_weekend_count = 0
                for e in evening_candidates:
                    if e in last_weekend_set:
                        if evening_last_weekend_count < 2:
                            evening_final.append(e)
                            evening_last_weekend_count += 1
                    else:
                        evening_final.append(e)
                    if len(evening_final) == evening_count:
                        break
                evening = evening_final
                used_this_week.update(evening)

                regular_count = len(editors) - morning_count - evening_count
                regular = get_sorted_candidates("常规班", used_this_week)[:regular_count]
                used_this_week.update(regular)

                if len(morning) < morning_count or len(evening) < evening_count or len(regular) < regular_count:
                    st.error(f"第 {week+1} 周排班失败：无法分配三班各 {morning_count}/{evening_count}/{regular_count} 人且不重叠。")
                    st.stop()

                for day in weekdays:
                    day_str = day.strftime("%Y-%m-%d")
                    schedule[day_str]["早班"] = morning
                    schedule[day_str]["晚班"] = evening
                    schedule[day_str]["常规班"] = regular
                    for e in morning:
                        shift_count[e]["早班"] += 1
                        shift_weeks[e]["早班"].add(week)
                    for e in evening:
                        shift_count[e]["晚班"] += 1
                        shift_weeks[e]["晚班"].add(week)
                    for e in regular:
                        shift_count[e]["常规班"] += 1
                        shift_weeks[e]["常规班"].add(week)

                # 优化后的周末值班分配逻辑
                never_assigned = [e for e in editors if shift_count[e]["周末值班"] == 0 and e not in exclude_weekend]
                weekend_assigned = []

                if len(never_assigned) >= weekend_count:
                    weekend_assigned = random.sample(never_assigned, weekend_count)
                else:
                    weekend_assigned = never_assigned.copy()
                    remaining = weekend_count - len(never_assigned)
                    eligible = [e for e in editors if e not in exclude_weekend and is_eligible_for_weekend(e, week) and e not in never_assigned]
                    eligible_sorted = sorted(eligible, key=lambda x: shift_count[x]["周末值班"])
                    weekend_assigned += eligible_sorted[:remaining]

                if len(weekend_assigned) < weekend_count:
                    st.error(f"第 {week+1} 周周末值班人选不足，请调整排除名单或減少排班周数。")
                    st.stop()

                for d in weekend:
                    day_str = d.strftime("%Y-%m-%d")
                    schedule[day_str]["周末值班"] = weekend_assigned
                    for e in weekend_assigned:
                        shift_count[e]["周末值班"] += 1
                        shift_weeks[e]["周末值班"].add(week)
                        past_weekend_weeks[e].add(week)

            calendar_df = []
            for date in sorted(schedule.keys()):
                for shift in schedule[date]:
                    for person in schedule[date][shift]:
                        calendar_df.append({"日期": date, "班次": shift, "姓名": person})
            calendar_df = pd.DataFrame(calendar_df)

            st.subheader("📊 排班表（姓名为列、日期为栏）")
            pivot_df = calendar_df.pivot(index="姓名", columns="日期", values="班次").fillna("")
            def sort_by_shift_group(df):
                shift_order = {"早班": 0, "晚班": 1, "常规班": 2, "周末值班": 3}
                def dominant_shift(row):
                    counts = defaultdict(int)
                    for val in row:
                        if val in shift_order:
                            counts[val] += 1
                    if counts:
                        return min(counts.items(), key=lambda x: (-x[1], shift_order.get(x[0], 99)))[0]
                    return ""
                df["主班次"] = df.apply(dominant_shift, axis=1)
                df_sorted = df.sort_values(by="主班次").drop(columns=["主班次"])
                return df_sorted
            sorted_pivot_df = sort_by_shift_group(pivot_df)
            def highlight_shifts(val):
                color_map = {
                    "早班": "background-color: #CCFFCC",
                    "晚班": "background-color: #ADD8E6",
                    "周末值班": "background-color: #FFFFCC"
                }
                return color_map.get(val, "")
            styled_df = sorted_pivot_df.style.applymap(highlight_shifts)
            st.dataframe(styled_df)
            pivot_csv = sorted_pivot_df.to_csv(index=True, encoding="utf-8-sig")
            st.download_button("📥 下载姓名为列的排班表 CSV", data=pivot_csv, file_name="排班表_姓名为列.csv", mime="text/csv")

            st.subheader("📥 下载彩色排班表 Excel（姓名为列、日期为栏）")
            wb = Workbook()
            ws = wb.active
            ws.title = "排班表"
            shift_colors = {
                "早班": "CCFFCC",
                "晚班": "ADD8E6",
                "常规班": "D3D3D3",
                "周末值班": "FFFFCC"
            }
            ws.append(["姓名"] + list(sorted_pivot_df.columns))
            for name, row in sorted_pivot_df.iterrows():
                row_data = [name]
                for date in sorted_pivot_df.columns:
                    shift = row[date]
                    row_data.append(shift)
                ws.append(row_data)
            for r in ws.iter_rows(min_row=2, min_col=2):
                for cell in r:
                    shift = cell.value
                    if shift in shift_colors:
                        fill = PatternFill(start_color=shift_colors[shift], end_color=shift_colors[shift], fill_type="solid")
                        cell.fill = fill
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            st.download_button(
                label="📥 下载彩色排班表 Excel",
                data=output,
                file_name="排班表_姓名为列_彩色.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.subheader("📊 每人排班统计")
            stat_df = pd.DataFrame.from_dict(shift_count, orient='index').reset_index()
            stat_df.columns = ["姓名", "早班", "晚班", "常规班", "周末值班"]
            st.dataframe(stat_df)
            csv = stat_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button("📥 下载彩色排班表 CSV", data=csv, file_name="排班统计表.csv", mime="text/csv")